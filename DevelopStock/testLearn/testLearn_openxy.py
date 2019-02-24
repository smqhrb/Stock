'''
https://blog.csdn.net/fengqiaoxian/article/details/81436840
'''
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# 创建一个engine='openpyxl'的 ExcelWriter 对象 writer
writer = pd.ExcelWriter(filename, engine='openpyxl')

try:
    # 加载指定的excel文件
    writer.book = load_workbook(filename)

    # 得到指定sheet的最后一行数据，因为是在原excel里面添加内容
    # 所以添加的信息应该从当前sheet最后一行的后面开始
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # 是否需要重新创建一下该sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:        
        idx = writer.book.sheetnames.index(sheet_name)        
        writer.book.remove(writer.book.worksheets[idx])        
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
except FileNotFoundError:
    # file does not exist yet, we will create it
    pass

if startrow is None:
    startrow = 0

df.apply(axis = 1)  #axis = 1 指定逐行添加，如果axis = 0，就是逐列添加信息
df.to_excel(writer, sheet_name, startrow=startrow, index=False, header=False)
# style set
font = Font(name='Calibri')
# 设置对齐样式
align = Alignment(horizontal='left', vertical='center')

# 得到需要修改样式的sheet
bigDataSheet = writer.book[sheet_name]

# 得到最大行数、列数
allRows = bigDataSheet.max_row
allColumns = bigDataSheet.max_column

#循环所有单元格，设置样式
for r in range(startrow, allRows+1):
    for c in range(1, allColumns):
        cellObj = bigDataSheet.cell(row = r, column= c)
        cellObj.font = font
        cellObj.alignment = align


writer.save()