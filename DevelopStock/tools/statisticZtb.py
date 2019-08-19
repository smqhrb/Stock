# 8个指标分别是
# 1.（A股收盘价低于-5%家数）-（跌停板家数），
# 2.A股最高板板数，
# 3.涨停板数量+涨停板被砸(最高价=涨停价，收盘价不等于涨停价)数量，
# 4.涨停板的（昨日涨停价-今日收盘价）/昨日涨停价的平均值，
# 5.涨停板被砸的（昨日涨停价-今日收盘价）/昨日涨停价的平均值，（最高价=涨停价，收盘价不等于涨停价）
# 6.1板的数量，
# 7.昨日1板的（昨日涨停价-今日收盘价）/昨日涨停价的平均值，
# 8.今日2板数量除以昨日1板数量

# 什么叫最高板 ：所有的涨停板里连板次数最多的（新股没开板的不计算在内）
# 这些数量里全是一字板的也丢掉

#using pysqlite 存储历史数据
#coding=utf-8
import getopt
import tushare as ts
import numpy as np
import pandas as pd
import os,time,sys,re,datetime
import urllib.request as urllib2
from bs4 import BeautifulSoup 
import random
from html.parser import HTMLParser
import time
import math
import json
import random
# import lxml.html
# from lxml import etree
# from pandas.io.html import read_html
# from pandas.compat import StringIO
# import xlwt
import time
 

import pandas as pd
from bs4 import BeautifulSoup

import openpyxl
import sqlite3 as db
class RandomHeader:
    def __init__(self):
        self.user_agent_list = [
            "Mozilla/5.0(Macintosh;IntelMacOSX10.6;rv:2.0.1)Gecko/20100101Firefox/4.0.1",
            "Mozilla/4.0(compatible;MSIE6.0;WindowsNT5.1)",
            "Opera/9.80(WindowsNT6.1;U;en)Presto/2.8.131Version/11.11",
            "Mozilla/5.0(Macintosh;IntelMacOSX10_7_0)AppleWebKit/535.11(KHTML,likeGecko)Chrome/17.0.963.56Safari/535.11",
            "Mozilla/4.0(compatible;MSIE7.0;WindowsNT5.1)",
            "Mozilla/4.0(compatible;MSIE7.0;WindowsNT5.1;Trident/4.0;SE2.XMetaSr1.0;SE2.XMetaSr1.0;.NETCLR2.0.50727;SE2.XMetaSr1.0)"
            "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/22.0.1207.1 Safari/537.1",
            "Mozilla/5.0 (X11; CrOS i686 2268.111.0) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.57 Safari/536.11",
            "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.6 (KHTML, like Gecko) Chrome/20.0.1092.0 Safari/536.6",
            "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.6 (KHTML, like Gecko) Chrome/20.0.1090.0 Safari/536.6",
            "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/19.77.34.5 Safari/537.1",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/536.5 (KHTML, like Gecko) Chrome/19.0.1084.9 Safari/536.5",
            "Mozilla/5.0 (Windows NT 6.0) AppleWebKit/536.5 (KHTML, like Gecko) Chrome/19.0.1084.36 Safari/536.5",
            "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3",
            "Mozilla/5.0 (Windows NT 5.1) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_8_0) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3",
            "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1062.0 Safari/536.3",
            "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1062.0 Safari/536.3",
            "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.1 Safari/536.3",
            "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.1 Safari/536.3",
            "Mozilla/5.0 (Windows NT 6.1) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.1 Safari/536.3",
            "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.0 Safari/536.3",
            "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/535.24 (KHTML, like Gecko) Chrome/19.0.1055.1 Safari/535.24"
        ]
    def GetHeader(self):
        headers={"User-Agent":random.choice(self.user_agent_list)}
        return headers
# from sqlalchemy import create_engine
# engine = create_engine(
#     "mysql+pymysql://root:root@localhost:3306/tushare?charset=utf8")
    
class dbOper:
    # https://www.cnblogs.com/lizunicon/p/4306730.htm
    # http://docs.jinkan.org/dodbOpercs/flask/patterns/sqlalchemy.html
    # https://www.cnblogs.com/zenan/p/10176767.html
    def __init__(self,db_file=""):
        # self.engine = create_engine('sqlite:///.\\%s'%db_file)
        self.cx = db.connect(db_file)
        
        # self.engine.connect(db_file)
    def cx_data(self,sql_select):
        cur =self.cx.cursor()
        cur.execute(sql_select)
        df =pd.DataFrame(cur.fetchall())
        cur.close()
        return df
    def update_data(self,sql_update):
        try:
            cur =self.cx.cursor()
            cur.execute(sql_update)
            cur.close()
            self.cx.commit()
        except Exception as e:
            self.cx.rollback()
            print(str(e))
    def pandas_tosql(self,df,tbl_name,ifexist):
        # df.to_sql(name = 'employee', con = con, if_exists='replace', index = None)
        df.to_sql(name = tbl_name, con = self.cx, if_exists=ifexist, index = None)
    def to_sql(self,dataf,tbl):
        '''
        dataf:存入数据的DataFrame
        tbl：数据库表名
        '''
        rowCnt =len(dataf)
        if(rowCnt==0):
            return
        # dataf.to_sql(name=tbl,con=self.cx,if_exists='append',index= False)
        colCnt =dataf.columns.tolist()
        cnt =len(colCnt)
        start = time.clock()

        # tuple(zip(df.Description, df.Code))
    

        try:
            self.cx.isolation_level = None
            self.cx.execute("BEGIN TRANSACTION")
            cur =self.cx.cursor()
            data_tuple= tuple(dataf.itertuples(index=False, name=None))
            cols =""
            vals =""
 
            for k in range(cnt):
                if(k==(cnt -1)):
                    cols = cols + colCnt[k]
                    vals = vals + "?"
                else:
                    cols = cols + "%s,"%(colCnt[k])
                    vals = vals + "?,"
            sql ="insert into %s(%s) values(%s);"%(tbl,cols,vals)
            cur.executemany(sql,data_tuple)
            # #very slow
            # for i in range(rowCnt):
            #     try: 
            #         # dataf.iloc[i:i+1].to_sql(name =tbl,  con=self.cx, if_exists='append',index= False)

            #         cols =""
            #         vals =""
            #         for k in range(cnt):
            #             if(k==(cnt -1)):
            #                 cols = cols + colCnt[k]
            #                 coltype =str(dataf[colCnt[k]].dtype)
            #                 if(coltype=="object"):
            #                     vals = vals + "\'%s\'"%(dataf.iloc[i:i+1][colCnt[k]].values[0])
            #                 else:
            #                     vals = vals + "%s"%(dataf.iloc[i:i+1][colCnt[k]].values[0])
            #             else:
            #                 cols = cols + "%s,"%(colCnt[k])
            #                 coltype =str(dataf[colCnt[k]].dtype)
            #                 if(coltype=="object"):
            #                     vals = vals + "\'%s\',"%(dataf.iloc[i:i+1][colCnt[k]].values[0])
            #                 else: 
            #                     vals = vals + "%s,"%(dataf.iloc[i:i+1][colCnt[k]].values[0])
            #         sql ="insert into %s(%s) values(%s);"%(tbl,cols,vals)
            #         # self.update_data(sql)
            #         cur.executemany(sql)
            #         #
            #         # self.cx.commit()
            #     except Exception as e:
            #         pass#
            #         # print('Error is ' + str(e))
            #         #self.con.rollback()  
            
            self.cx.execute("COMMIT") 
            
            end = time.clock()
            print("...写数据库耗时:%.4f   存入数据行数=%s"%(end-start,rowCnt))
        except Exception as e:
            # cur.close()
            # cur.execute("rollback")
            print(str(e))


class dataOper:
    '''
    fetch_hist_data
    按日期 排序
    '''
    def __init__(self,db_file=""):
        self.db =dbOper(db_file)
    # 1.GetStockList
    def GetStockList(self):
        '''
        1.get all count
        2.every request 20 or 40 or 80
        3.get data
        '''
        everyPage =40
        print("------开始读取网上股票列表信息")   
        #get stock count
        # http://vip.stock.finance.sina.com.cn/mkt/#hs_a
        # http://vip.stock.finance.sina.com.cn/quotes_service/api/json_v2.php/Market_Center.getHQNodeStockCount?node=hs_a 
        # 40 
        # http://vip.stock.finance.sina.com.cn/quotes_service/api/json_v2.php/Market_Center.getHQNodeData?page=1&num=40&sort=symbol&asc=1&node=hs_a&symbol=&_s_r_a=init
        getH =RandomHeader()
        headers =getH.GetHeader()
        url ="http://vip.stock.finance.sina.com.cn/quotes_service/api/json_v2.php/Market_Center.getHQNodeStockCount?node=hs_a"
        req = urllib2.Request(url, headers = headers)
        content = urllib2.urlopen(req).read()
        text =content.decode('utf8')
        # print(text)
        count =text[(text.find('("')+2):text.find('")')]
        # print(count)
        iCount =int(count)
        pageCount =math.ceil(iCount/everyPage)
        df =pd.DataFrame()
        for page in range(1,pageCount+1):
            data_url =("http://vip.stock.finance.sina.com.cn/quotes_service/api/json_v2.php/Market_Center.getHQNodeData?page=%s&num=%s&sort=symbol&asc=1&node=hs_a&symbol=&_s_r_a=init")%(page,everyPage)
            getH =RandomHeader()
            headers =getH.GetHeader()
            req = urllib2.Request(data_url, headers = headers)
            content = urllib2.urlopen(req).read()
            text =content.decode('GBK')
            text =self.jsonfy(text)# to dict 不带双引号的json的key标准化
            if(len(df)==0):
                df =pd.DataFrame(data=text)
            else:
                df1 =pd.DataFrame(data=text)
                df1['sssj'] =""
                len_df1 = len(df1)
                for i in range(len_df1):
                    sssj =self.GetGpFxrq(df1['code'][i])
                    df1['sssj'][i]=sssj
                    time.sleep(random.random()+1)
                r = df1[['code','name','sssj']]
                self.db.to_sql(r,'stock_code_name')
            time.sleep(random.random()+1)

        print("------结束读取网上股票列表信息")   
    def jsonfy(self,s):
        #此函数将不带双引号的json的key标准化
        obj = eval(s, type('js', (dict,), dict(__getitem__=lambda s, n: n))())
        return obj 
    # 2.GetStockData open close pre_close pct_chg涨跌幅度
    def GetStockData(self):
        pro = ts.pro_api()
        i=0
        sql ="select code,sssj from stock_code_name where name not like \'%ST%\'"
        df_sssj =self.db.cx_data(sql)
        len_df =len(df_sssj)
        for k in range(len_df):
            #ts must delay 300ms
            time.sleep(0.3)
            #1 最后日期
            code =df_sssj[0][k]
            if code>='600000':
                code ="%s.SH"%code
            else:
                code ="%s.SZ"%code
            df_date =self.db.cx_data("select max(trade_date) as tdate from stock_day where ts_code =\'%s\'"%code)
            if(df_date[0][0]==None):
                start =df_sssj[1][k]

                if(start==""):
                    start ='20000101'
                else:
                    dateTime_p = datetime.datetime.strptime(start,'%Y-%m-%d')
                    start =datetime.datetime.strftime(dateTime_p,'%Y%m%d')

            else:
                start =df_date[0][0]
                start =str(start)
                dateTime_p = datetime.datetime.strptime(start,'%Y%m%d')
                detaday=datetime.timedelta(days=1)
                da_days=dateTime_p+detaday
                start =datetime.datetime.strftime(da_days,'%Y%m%d')

            now_time = datetime.datetime.now()#现在
            end =now_time.strftime('%Y%m%d')
            #2 重最后日期 到 最新日期的股价
            if(start>end):
                print("...开始日期%s>结束日期%s,数据库中已经是最新数据 ..."%(start,end))
                break
            df1 = pro.daily(ts_code=code, start_date=start, end_date=end)
            df=df1.sort_values(by=['trade_date'])
            df['high_pct_change'] =(df['high'] -df['pre_close'])*100/df['pre_close']
            df=df.reset_index()
            df.drop(['index'],axis=1,inplace=True)
            #3 插入数据库
            self.db.to_sql(df,'stock_day')
            print("...[%d]finish writing Stock =%s to table stock_day"%(k,code))
    # 4.GetGpFxrq
    
    def GetGpFxrq(self,Code):
        # http://stockpage.10jqka.com.cn/000600/company/
        getH =RandomHeader()
        headers =getH.GetHeader()
        url ="http://basic.10jqka.com.cn/%s/company.html#stockpage"%Code
        req = urllib2.Request(url, headers = headers)
        content = urllib2.urlopen(req).read()

        soup = BeautifulSoup(content,features="lxml")

        date =""
        try:
            div0 = soup.find("div",{"class":"m_box company_detail"})
            # div0 = soup.find("div",{"id":"publish"})
            dbrow =div0.findAll("tr")
            
            dbrow_cnt =len(dbrow)
            
            if(dbrow_cnt>1):
                date =dbrow[1].findAll("td")[0].contents[1].text
        except Exception as e:
            pass
        return date
        # pass
    def AnalysisIndex(self,day_rq=""):
        wdCheck =WorkDay(day_rq)
        days =wdCheck.get_recent_twotradeday(3)
        self.AnalysisIndexDay(days)

    def AnalysisIndexDay(self,days):
        '''
        # 5.Analysis data 
        # 1.（A股收盘价低于-5%家数）,（跌停板家数），
        # 2.A股最高板板数，
        # 3.涨停板数量,涨停板被砸,涨停板数量+涨停板被砸(最高价=涨停价，收盘价不等于涨停价)数量，
        # 4.涨停板的（今日收盘价 -昨日涨停价）/昨日涨停价的平均值，
        # 5.涨停板,涨停板被砸的（今日收盘价-昨日涨停价）/昨日涨停价的平均值，（最高价=涨停价，收盘价不等于涨停价）
        # 6.1板的数量，
        # 7.昨日1板的（昨日涨停价-今日收盘价）/昨日涨停价的平均值，
        # 8.今日2板数量除以昨日1板数量

        # 什么叫最高板 ：所有的涨停板里连板次数最多的（新股没开板的不计算在内）
        # 这些数量里全是一字板的也丢掉
        '''
        ZTB_THRESOLD =9.9
        # 查找交易日 和 上一交易日
        previous_day =days[1]
        first_day =days[0]
        # 删除已经存在的记录
        sql ="delete from stock_ztb where day ='%s'"%first_day
        self.db.update_data(sql)
        #
        ssrq = self.db.cx_data("select strftime('%Y%m%d',sssj) sssj from stock_code_name")
        #1 A股收盘价低于-5%家数）,（跌停板家数）
        sql ="select count(*) from stock_day where pct_chg<-5  and trade_date ='%s'"%first_day #低于-5%家数
        # 'select count(*) from stock_day where pct_chg<-5 and trade_date='20190816';'
        pd1 =self.db.cx_data(sql)
        lowNeg5 =pd1[0][0]
        # 

        sql ="select count(*) from stock_day where pct_chg<-9.8 and high!=low and trade_date ='%s'"%first_day #跌停板家数
        pd1 =self.db.cx_data(sql)
        dtbCnt =pd1[0][0]
        # 'select count(*) from stock_day where pct_chg<-9.8 and high!=low and trade_date ='20190816';'

        #2 A股最高板板数:从最后交易日涨停板,往前计算得到的最多股票板数
        # "select ts_code,trade_date,pct_chg from stock_day where pct_chg>9.9 and ts_code = "#A股最高板板数
        sql = "select ts_code from stock_day where pct_chg>%s and trade_date ='%s'"%(ZTB_THRESOLD,first_day)
        pd_cx =self.db.cx_data(sql)
        cnt_pd =len(pd_cx)
       
        dateTime_p =datetime.datetime.today()
        detaday=datetime.timedelta(days=730)
        da_days=dateTime_p-detaday
        start =datetime.datetime.strftime(da_days,'%Y%m%d') 
        ts_list =[]  
        ts_code_new =[]      
        for i in range(cnt_pd):
            ccontent =[]
            ts_code = pd_cx[0][i]
            sql ="select pct_chg,trade_date from stock_day where ts_code ='%s' and trade_date>'%s' order by trade_date desc"%(ts_code,start)
            pd_ts = self.db.cx_data(sql)
            pd_ts_cnt =len(pd_ts)
            ztb_cnt =0
            flag =0
            
            for k in range(pd_ts_cnt):
                pct_chg =pd_ts[0][k]
                # 查找 上市 没开板的股票代码
                check =ssrq[ssrq[0]==str(pd_ts[1][k])]
                if(len(check)>0):
                    flag =1
                    ts_code_new.append(ts_code)
                    break
                if(pct_chg<ZTB_THRESOLD):
                    break
                else:
                    ztb_cnt =ztb_cnt +1
            if(flag==1):
                continue
            ccontent.append(ts_code)
            ccontent.append(ztb_cnt)
            ts_list.append(ccontent)
        df =pd.DataFrame(ts_list,columns=['ts_code','ztb_cnt'])
        maxZtbCnt = df['ztb_cnt'].max()
        # loop to search A股最高板板数

        #3 涨停板数量,涨停板被砸,涨停板数量+涨停板被砸(最高价=涨停价，收盘价不等于涨停价)数量 
        sql = "select count(*) from stock_day where pct_chg>%s and high!=low and trade_date ='%s'"%(ZTB_THRESOLD,first_day) #涨停板数量
        pd_3 =self.db.cx_data(sql)
        ztbCnt =pd_3[0][0]
        # 'select count(*) from stock_day where pct_chg>9.9 and high!=low and trade_date ='20190816';'

        sql = "select count(*) from stock_day where ((high-pre_close)*100/pre_close)>%s and high!=close and trade_date ='%s'"%(ZTB_THRESOLD,first_day)#涨停板被砸数量
        pd_3 =self.db.cx_data(sql)
        ztbBz =pd_3[0][0]
        # 'select count(*) from stock_day where ((high-pre_close)*100/pre_close)>9.9 and high!=close and trade_date = '20190816';'
        
        #4 昨日涨停板的（今日收盘价 -昨日涨停价）/昨日涨停价的平均值
        sql ="select (close - pre_close)*100/pre_close from stock_day where trade_date='%s' and ts_code in(select ts_code  from stock_day where pct_chg>%s and trade_date='%s')"%(first_day,ZTB_THRESOLD,previous_day)
        pd_4 =self.db.cx_data(sql)
        ztbPriceAvg ='%.2f'%(pd_4[0].mean()) 
        #5 昨日涨停板被砸的（今日收盘价-昨日涨停价）/昨日涨停价的平均值
        sql = "select (close - pre_close)*100/pre_close from stock_day where  trade_date='%s' and ts_code in(select ts_code from stock_day where high_pct_change>%s and close!=high and trade_date='%s')"%(first_day,ZTB_THRESOLD,previous_day)
        pd_5 =self.db.cx_data(sql)
        ztbBzPriceAvg ='%.2f'%(pd_5[0].mean()) 
        #6 1板的数量
        sql ="select count(*) from stock_day where pct_chg>%s and trade_date='%s' and ts_code in (select ts_code from stock_day where pct_chg<%s and trade_date='%s')"%(ZTB_THRESOLD,first_day,ZTB_THRESOLD,previous_day)
        pd_6 =self.db.cx_data(sql)
        oneZtCnt =pd_6[0][0]         
        #7 昨日1板的（昨日涨停价-今日收盘价）/昨日涨停价的平均值
        # sql ="select (close - pre_close)*100/pre_close from stock_day where  trade_date='%s' and high!=close and ts_code in (select ts_code from stock_day where pct_chg>%s and trade_date='%s')"%(first_day,ZTB_THRESOLD,previous_day)
        sql ="select (close - pre_close)*100/pre_close from stock_day where pct_chg<%s and trade_date='%s' and  ts_code in (select ts_code from stock_day where pct_chg>%s and trade_date='%s') and  ts_code in (select ts_code from stock_day where pct_chg<%s and trade_date='%s')"%(ZTB_THRESOLD,days[0],ZTB_THRESOLD,days[1],ZTB_THRESOLD,days[2])#昨日1板数量
        pd_7 =self.db.cx_data(sql)
        zrOneZtAvg ='%.2f'%(pd_7[0].mean())          
        #8 今日2板数量除以昨日1板数量
        
        sql ="select count(*) from stock_day where pct_chg>%s and trade_date='%s' and  ts_code in (select ts_code from stock_day where pct_chg>%s and trade_date='%s') and  ts_code in (select ts_code from stock_day where pct_chg<%s and trade_date='%s')"%(ZTB_THRESOLD,days[0],ZTB_THRESOLD,days[1],ZTB_THRESOLD,days[2])#今日2板数量
        pd_8 =self.db.cx_data(sql)
        twoZtb =pd_8[0][0] 

        sql ="select count(*) from stock_day where pct_chg<%s and trade_date='%s' and  ts_code in (select ts_code from stock_day where pct_chg>%s and trade_date='%s') and  ts_code in (select ts_code from stock_day where pct_chg<%s and trade_date='%s')"%(ZTB_THRESOLD,days[0],ZTB_THRESOLD,days[1],ZTB_THRESOLD,days[2])#昨日1板数量
        pd_8 =self.db.cx_data(sql)
        oneZtb =pd_8[0][0] 
        if(oneZtb==0):
            today2Div1 =0
        else:
            today2Div1 ='%.2f'%(twoZtb/oneZtb)
        # update to xlsx
        dateTime_p = datetime.datetime.strptime(first_day,'%Y%m%d')
        wdCheck = WorkDay()
        week_str = wdCheck.get_week_day(dateTime_p)
        sql ="insert into stock_ztb(\
        day,week_str,lowNeg5,dtbCnt,maxZtbCnt,ztbCnt,ztbBz,ztbPriceAvg,ztbBzPriceAvg,oneZtCnt,zrOneZtAvg,today2Div1)\
        values('%s','%s',%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"%(first_day,week_str,lowNeg5,dtbCnt,maxZtbCnt,ztbCnt,ztbBz,ztbPriceAvg,ztbBzPriceAvg,oneZtCnt,zrOneZtAvg,today2Div1)
        self.db.update_data(sql)

        # pass
    def writeXls(self,file_name):
       
        #加载文件
        wb = openpyxl.load_workbook(file_name)
        #获得sheet名称
        sheetNames = wb.sheetnames
        print(sheetNames)
        #sheetName1 = sheetNames[0]
        #根据名称获取第一个sheet
        #sheet1 = wb[sheetName1]
        #根据索引获得第一个sheet
        sheet1 = wb.worksheets[0]
        
        # L = ['张三', '李四', '王五']
        # #excel中单元格为B2开始，即第2列，第2行
        # for i in range(len(L)):
        #     sheet1.cell(i+2, 2).value=L[i]
        offset =2
        day =0 
        week_str =1 
        lowNeg5 =2 
        dtbCnt =3 
        maxZtbCnt =4 
        ztbCnt =5 
        ztbBz =6 
        ztbPriceAvg =7 
        ztbBzPriceAvg =8 
        oneZtCnt =9 
        zrOneZtAvg =10 
        today2Div1 =11 
        sql ="select day,week_str,lowNeg5,dtbCnt,maxZtbCnt,ztbCnt,ztbBz,ztbPriceAvg,ztbBzPriceAvg,oneZtCnt,zrOneZtAvg,today2Div1 from stock_ztb order by day desc"
        pd_ztb =self.db.cx_data(sql)
        pd_ztb_len =len(pd_ztb)
        for i in range(pd_ztb_len):
            lineOffset =2
            sheet1.cell(i+lineOffset,day + offset).value=pd_ztb[day][i]
            sheet1.cell(i+lineOffset,week_str+ offset).value=pd_ztb[week_str][i]
            sheet1.cell(i+lineOffset,lowNeg5+ offset).value=pd_ztb[lowNeg5][i]
            sheet1.cell(i+lineOffset,dtbCnt+ offset).value=pd_ztb[dtbCnt][i]
            sheet1.cell(i+lineOffset,maxZtbCnt+ offset).value=pd_ztb[maxZtbCnt][i]
            sheet1.cell(i+lineOffset,ztbCnt+ offset).value=pd_ztb[ztbCnt][i]
            sheet1.cell(i+lineOffset,ztbBz+ offset).value=pd_ztb[ztbBz][i]
            sheet1.cell(i+lineOffset,ztbPriceAvg+ offset).value=pd_ztb[ztbCnt][i] + pd_ztb[ztbBz][i]
            sheet1.cell(i+lineOffset,ztbPriceAvg+ offset+1).value=pd_ztb[ztbPriceAvg][i]
            sheet1.cell(i+lineOffset,ztbBzPriceAvg+ offset+1).value=pd_ztb[ztbBzPriceAvg][i]
            sheet1.cell(i+lineOffset,oneZtCnt+ offset+1).value=pd_ztb[oneZtCnt][i]
            sheet1.cell(i+lineOffset,zrOneZtAvg+ offset+1).value=pd_ztb[zrOneZtAvg][i]
            sheet1.cell(i+lineOffset,today2Div1+ offset+1).value=pd_ztb[today2Div1][i]
        #保存数据，如果提示权限错误，需要关闭打开的excel

        wb.save(file_name)

class WorkDay:
    def __init__(self,start=""):
        if(start==""):
            dateTime_p =datetime.datetime.today()
            self.dateTime_t =dateTime_p -datetime.timedelta(hours=16)
        else:
            self.dateTime_t = datetime.datetime.strptime(start,'%Y%m%d')

    def get_day_type(self,query_date):
        getH =RandomHeader()
        headers =getH.GetHeader()
        time.sleep(1)
        url = 'http://tool.bitefu.net/jiari/?d=' + query_date
       
        req = urllib2.Request(url, headers = headers)
        content = urllib2.urlopen(req).read()
        text =content.decode('utf8')
        if text == '"0"':
            return 0
        elif text == '1':
            return 1
        elif text == '2':
            return 2
        # 上面的url接口  工作日对应结果为 0, 休息日对应结果为 1, 节假日对应的结果为 2

    def today_is_tradeday(self):
        query_date = datetime.datetime.strftime(datetime.datetime.today(), '%Y%m%d')
        print(query_date)
        return self.get_day_type(query_date)

    def get_recent_twotradeday(self,day_cnt=2):
        days =[]
        i =0
        while(len(days)<day_cnt):
            # dateTime_p =datetime.datetime.today()
            # dateTime_t =dateTime_p -datetime.timedelta(hours=16)
            # #dateTime_p = datetime.datetime.strptime(start,'%Y%m%d')
            detaday=datetime.timedelta(days=i)
            da_days=self.dateTime_t-detaday
            start =datetime.datetime.strftime(da_days,'%Y%m%d')
            day_type =self.get_day_type(start)
            if(day_type==0):
                days.append(start)
            i =i+1

        return days

    def get_week_day(self,date):

        week_day_dict = {
            0 : '星期一',
            1 : '星期二',
            2 : '星期三',
            3 : '星期四',
            4 : '星期五',
            5 : '星期六',
            6 : '星期天',
        }
        day = date.weekday()
        return week_day_dict[day]

def main():
    print(sys.argv[0])
    print(sys.argv[1])
    print(sys.argv[2])
    
    path =sys.argv[1]
    db_name =path+"\\stockA.db"
    xlsTest =dataOper(db_name)
    wk_type =sys.argv[2]

    if(wk_type=='stock'):# stock's code and name
        xlsTest.GetStockList()
    else:
        
        # xlsTest.GetStockData()
        #caculate index data
        
        day_start =sys.argv[3]
        
        day_end =sys.argv[4]
        # day_end_p =datetime.datetime.strptime(day_end,'%Y%m%d')
        while(day_start<=day_end):
            print('...更新指标的的日期为%s'%day_start)
            xlsTest.AnalysisIndex(day_start)
            day_start_p =datetime.datetime.strptime(day_start,'%Y%m%d')
            detaday =datetime.timedelta(days=1)
            da_days =day_start_p + detaday
            day_start =datetime.datetime.strftime(da_days,'%Y%m%d')

        xlsTest.writeXls(path+"\\stockA.xlsx")
        print('...更新指标写入%s'%(path+"\\stockA.xlsx"))

if __name__ == '__main__':
    main()
    # https://github.com/sadjjk/tonghuashun_industry
    # https://www.jianshu.com/p/13381aac9245
    # test =dataOper("stockA.db")
    # # test.GetStockListFromSina()
    # # test.GetGpFxrq("300722")
    # # test.GetStockData()
    # test.AnalysisIndex('20190814')
    # test.writeXls("stockA.xlsx")


# http://basic.10jqka.com.cn/000600/company.html#stockpage

